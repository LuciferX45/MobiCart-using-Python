import tkinter as tk
from openpyxl import *
from tkinter import *  
from tkinter import messagebox
from Cart import ShoppingCart
import LinkedList as LL
from datetime import *

#loads and activates the excel sheet
wb = load_workbook("D:\I034, I042, I042, I049_DSA Project_MobiCart\Details.xlsx")
sheet = wb.active

#To create the excel sheet with the columns 
def headings():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 50
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 50

    #the headings of the columns
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Contact"
    sheet.cell(row=1, column=3).value = "Email Id"
    sheet.cell(row=1, column=4).value = "Payment Method"
    sheet.cell(row=1, column=5).value = "Amount"
    sheet.cell(row=1, column=6).value = "Items"
    sheet.cell(row=1, column=7).value = "Delivery Type"
    sheet.cell(row=1, column=8).value = "Address"

    #saves the changes
    wb.save("D:\I034, I042, I042, I049_DSA Project_MobiCart\Details.xlsx")


#the Store GUI
def viewStore(lPhone,lPrice):
    #outline frame 
    global storeWindow 
    storeLabelFrame = LabelFrame(storeWindow, text="Store Items")
    storeLabelFrame.configure(bg = "light green")
    storeLabelFrame.pack(fill="both", expand="yes", padx="50", pady="20")

    #creates item frame
    storeItemsFrame = Frame(storeLabelFrame)
    storeItemsFrame.pack(padx="20", pady="10")
    
    #linked list for store items
    store = LL.LinkedList()
    for i in range(len(lPhone)):
        store.addNode(lPhone[i])
        store.addNode(lPrice[i])

    size = store.size()
    
    #creates the scrollable frame
    canvas = Canvas(storeItemsFrame,width=400, height=350)
    scroll_bar = Scrollbar(storeItemsFrame,orient= "vertical",command=canvas.yview)
    scroll_bar.pack(side=RIGHT,fill=Y)

    scrollable_frame = Frame(canvas)
    scrollable_frame.bind("<Configure>",
                    lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scroll_bar.set)
    canvas.pack(side=LEFT, fill=BOTH, expand=True)


    #displays the items in the frame
    for i in range(2,size-1,2):

        itemFrame = Frame(scrollable_frame,  pady="5")
        itemFrame.pack(fill="both", expand="yes")

        #gets phone and price from Linked List
        ph = store.Ndisplay(i)
        pr = store.Ndisplay(i+1)

        #creates labels to fit in the frame
        nameLabel = Label(itemFrame, text=ph,font=("Bahnschrift SemiBold",11),fg="blue")
        priceLabel = Label(itemFrame, text="₹ %s"%pr,font=("Bahnschrift SemiBold",13),fg="red")  
        addToCartBtn = Button(itemFrame, text="Add To Cart",cursor="hand2", command=lambda e=[ph,pr]: addItemToCart(e)) 
        btnImage=PhotoImage(file="images/addToCart.png")       
        addToCartBtn.image= btnImage 
        addToCartBtn.config(image=btnImage,width="40",height="30")

        nameLabel.pack(side="left")
        priceLabel.pack(side="left",fill="both", expand="yes" )
        addToCartBtn.pack(side="right" )

    #button to access the cart    
    btnGoCart = Button(storeWindow, text="Go To Cart", font=("Bahnschrift SemiBold",13),fg="red",bg="white",cursor="hand2", command=viewCart )
    btnGoCart.pack(pady="6")

#the Cart GUI
def viewCart():  
    #outline frame 
    cartWindow = Toplevel()
    cartWindow.title("The Cart")
    cartWindow.grab_set()
    global cart
    cartItems = cart.getCartItems()

    #items frame
    cartItemsLabelFrame = LabelFrame(cartWindow,text="Cart Items")
    cartItemsLabelFrame.pack(fill="both", expand="yes", padx="50", pady="30")

    cartItemsFrame = Frame(cartItemsLabelFrame, padx=3, pady=3)
    cartItemsFrame.pack()
    index = 0
    
    #displays the cart items
    for item in cartItems:
        itemFrame = Frame(cartItemsFrame,  pady="5")
        itemFrame.pack(fill="both", expand="yes")

        #labels to fit in the frame
        nameLabel = Label(itemFrame, text=item[0],font=("Bahnschrift SemiBold",13),fg="blue")
        priceLabel = Label(itemFrame, text="₹ %s"%item[1],font=("Bahnschrift SemiBold",13),fg="red")  
        remToCartBtn = Button(itemFrame, text="Remove From Cart", font=("Bahnschrift SemiBold",11),fg="red",bg="white",cursor="hand2", command=lambda i=index: removeFromCart(i,cartWindow) )

        nameLabel.pack(side="left")
        priceLabel.pack(side="left")
        remToCartBtn.pack(side="right" ) #button to remove from cart
        index += 1

    #creates checkout frame
    checkOutFrame = Frame(cartWindow, pady="10")
    AMOUNT.set(cart.getTotalPrice())
    totalPriceLabel = Label(checkOutFrame, text="Total Price : ₹ %s" % cart.getTotalPrice(), font=("Bahnschrift SemiBold",14),fg="indigo")
    totalPriceLabel.pack(side="left")
    buyBtn = Button(checkOutFrame, text="Buy Now", font=("Bahnschrift SemiBold",15),fg="indigo",bg="white",cursor="hand2", command=lambda:[buyCommand(cartWindow),PersonalDetails(cartWindow)])
    buyBtn.pack(side="left",padx="10") #button to buy 
    checkOutFrame.pack()

    #button to go back to store
    backToStoreBtn = Button(cartWindow, text="Back To Store", font=("Bahnschrift SemiBold",15),fg="red",bg="white",cursor="hand2",command=cartWindow.destroy)
    backToStoreBtn.pack(pady="6")

    cartWindow.mainloop()

#function for adding to cart
def addItemToCart(item=None):
    global cart
    cart.addToCart(item)
    messagebox.showinfo(title="Success", message="Item %s Added To The Cart !!"%item[0] )

#function for removing from cart
def removeFromCart(itemIndex=None,cartWindow=None):
    global cart
    cart.removeFromCart(itemIndex)
    messagebox.showinfo(title="success",message="Item Removed")
    cartWindow.destroy()
    viewCart()

#function to display successful purchase
def buyCommand(cartWindow):
    global cart
    cartWindow.destroy()    
    messagebox.showinfo(title="Success",message="Proceeding to Checkout")

#function to destory the GUI 
def dest(widget):
    widget.destroy()

#Personal Details GUI
def PersonalDetails(cartWindow):
    #creates the outline
    cartWindow.destroy()
    root = Tk()
    root.configure(background='light green')
    root.title("Personal Details")
    root.geometry("500x300")

    #creates the labels
    heading = Label(root, text="Form",font=("Bahnschrift SemiBold",15), bg="light green")
    name = Label(root, text="Name",font=("Bahnschrift SemiBold",12), bg="light green")
    contact_no = Label(root, text="Contact No.",font=("Bahnschrift SemiBold",12), bg="light green")
    email_id = Label(root, text="Email ID",font=("Bahnschrift SemiBold",12), bg="light green")

    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    contact_no.grid(row=2, column=0)
    email_id.grid(row=3, column=0)

    #creates the textbox to take data
    name_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)

    name_field.bind("<Return>", lambda e: contact_no_field.focus_set())
    contact_no_field.bind("<Return>", lambda e: email_id_field.focus_set())
    email_id_field.bind("<Return>")

    name_field.grid(row=1, column=1, ipadx="100")
    contact_no_field.grid(row=2, column=1, ipadx="100")
    email_id_field.grid(row=3, column=1, ipadx="100")

    #button to submit and move to next window
    submit = Button(root, text="Submit and Next", fg="Black",bg="Red", 
            command= lambda: [NAME.set(name_field.get()),CONTACT.set(contact_no_field.get()),
            EMAIL.set(email_id_field.get()),dest(root),delivery()])
    submit.grid(row=5, column=1,pady=50)
    
    root.mainloop()

#Delivery GUI
def delivery():
    #creates the outline
    root = Tk()
    root.title('Delivery Details')
    root.geometry('450x200')
    root.configure(bg = "light yellow")

    #creates labels and textbox to display and take data
    name = Label(root, text="Name")
    name.grid(row=0, pady=2)
    name_field = Label(root,text=NAME.get(),width=30)
    name_field.grid(row=0, column=1, pady=2)

    contact = Label(root, text="Contact")
    contact.grid(row=1, pady=2)
    contact_field = Label(root,text=CONTACT.get() ,width=10)
    contact_field.grid(row=1, column=1, pady=2)

    address = Label(root, text="Delivery Address")
    address.grid(row=2, pady=2)
    address_field = Entry(root,width=50)
    address_field.grid(row=2, column=1, pady=2)

    type = Label(root, text="Choose type of Delivery")
    type.grid(row=5, columnspan=2, pady=2)

    v = IntVar(value=0)

    #creates option list
    Radiobutton(root, text='Standard Delivery', variable=v, value=1,
                command=DELIVERY.set("Standard Delivery")).grid(row=6, columnspan=2, pady=2)
    Radiobutton(root, text='Prime Delivery', variable=v, value=2,
                command=DELIVERY.set("Prime Delivery")).grid(row=7, columnspan=2, pady=2)

    #submit button
    submit =Button(root, text='Submit', width=25,
            command= lambda: [ADDRESS.set(address_field.get()),dest(root),Payment()])
    submit.grid(row=10, columnspan=2, pady=2)
    
    root.mainloop()

#Payment GUI
def Payment():
    #creates the outline
    root = Tk()
    root.configure(background='light blue')
    root.title("Payment")
    root.geometry("400x300")
	
    #creates the labels
    name = Label(root, text="Name")
    amount = Label(root, text="Amount")
    payment_option = Label(root, text="Payment Option")
    
    name.grid(row=1, column=0, pady = 5)
    amount.grid(row=2, column=0, pady = 5)
    payment_option.grid(row=3, column=5)

    #creates textbox to display or take data
    name_field = Label(root, text=NAME.get(),bg = "white")
    name_field.grid(row=1,column=1,padx=5)
    amount_field = Label(root, text=AMOUNT.get(),bg="white")
    amount_field.grid(row=2,column=1,padx=5)

    
    v = IntVar(value=0)

    #creates the options list
    Radiobutton(root, text = "UPI Payment",variable = v, value = 1, 
                command= lambda: [PAYMENT.set("UPI"),UPI()] ).grid(row = 5, column = 5)
    Radiobutton(root, text = "Credit/Debit Card",variable = v, value = 2,
                command= lambda:[PAYMENT.set("Credit/Debit Card"),CC()]).grid(row = 6, column = 5)
    Radiobutton(root, text = "Cash on Delivery",variable = v, value=3,
                command=PAYMENT.set("Cash on Delivery")).grid(row = 7, column = 5)

    #button to submit and move to next window
    submit = Button(root, text="Submit and Next", fg="Black",bg="Red"
                    ,command= lambda:[dest(root),Billing()])
    submit.grid(row=10, column=5,pady = 50)

    root.mainloop()

#Upi GUI
def UPI():
    #creates the outline
	upi = Toplevel()
	upi.title("UPI Payment")
	upi.geometry("400x300")
	upi.configure(bg = "light blue")

    #creates labels and textbox to display and take data
	id = Label(upi,text="UPI ID")
	id.grid(row=0, column=1,pady= 50)
	id_field = Entry(upi)
	id_field.grid(row=0, column = 2, padx=30)
	pin = Label(upi,text="UPI PIN")
	pin.grid(row=1,column=1)
	pin_field = Entry(upi)
	pin_field.grid(row=1, column = 2, padx=30)

    #submit button
	submit = Button(upi,text = "Submit",bg = "red",command=upi.destroy)
	submit.grid(row = 2,column= 2,pady=20)

	upi.mainloop()

#credit card GUI
def CC():
    #creates outline
	cc = Toplevel()
	cc.title("Credit Card/Debit Card")
	cc.geometry("400x300")
	cc.configure(bg = "light green")

    #creates labels and textbox to display and take data
	chn = Label(cc,text= "Cardholder Name")
	chn.grid(row=0,column=1,pady=5)
	cn = Label(cc,text="Card No")
	cn.grid(row=1,column=1,pady=5)
	exp = Label(cc,text="Expiry Date")
	exp.grid(row=2,column=1,pady=5)
	cvv = Label(cc,text="CVV")
	cvv.grid(row=3,column=1,pady=5)

	chn_field = Entry(cc)
	chn_field.grid(row=0, column = 2, padx=30)
	cn_field = Entry(cc)
	cn_field.grid(row=1, column = 2, padx=30)
	exp_field = Entry(cc)
	exp_field.grid(row=2, column = 2, padx=30)
	cvv_field = Entry(cc)
	cvv_field.grid(row=3, column = 2, padx=30)

    #submit button
	submit = Button(cc,text = "Submit",bg = "red",command= cc.destroy)
	submit.grid(row = 4,column= 2,pady=10)

	cc.mainloop()	

def Billing():
    #creates the outline
    root = Tk()
    root.title('Billing')
    root.geometry('500x500')
    root.configure(bg = "White")

    #creates the labels and textbox to display and take data
    heading = Label(root, text="INVOICE",font=("Bahnschrift SemiBold",15), bg="light green")
    heading.pack(side=TOP)

    today = date.today()
    d1 = today.strftime("%d/%m/%Y")

    curr_date = Label(root, text=d1,font=("Bahnschrift SemiBold",12))
    curr_date.pack(side=TOP,pady=20)

    name = Label(root, text="Customer Name: "+ NAME.get(),font=("Bahnschrift SemiBold",12))
    name.pack(side=TOP)

    address = Label(root, text="Delivery Address: "+ ADDRESS.get(),font=("Bahnschrift SemiBold",12))
    address.pack(side=TOP)

    delivery = Label(root, text="Delivery Type: "+ DELIVERY.get(),font=("Bahnschrift SemiBold",12))
    delivery.pack(side=TOP)

    #changes date of delivery based on type of delivery
    if DELIVERY.get() == "Standard Delivery":
        dd = date.today() + timedelta(days = 7)
    else:
        dd = date.today() + timedelta(days = 2)

    d2 = dd.strftime("%d/%m/%Y")
    
    deli_date = Label(root, text=d2,font=("Bahnschrift SemiBold",12))
    deli_date.pack(side=TOP)

    item_heading = Label(root,text="Items" + "\t\t\t\t" + "Price",font=("Bahnschrift SemiBold",12))
    item_heading.pack(side=TOP,pady=20)

    #creates a string of the item for displaying in invoice
    itemslist = ""
    c = cart.getCartItems()
    for i in c:
        items = Label(root,font=("Bahnschrift SemiBold",12))
        items.config(text=i[0] + '\t\t\t' +"₹"+str(i[1]))
        items.pack(side=TOP)
        itemslist += i[0] + ","

    ITEMS.set(itemslist)

    total = Label(root,text="TOTAL= ₹"+AMOUNT.get(),font=("Bahnschrift SemiBold",12), fg = "red")
    total.pack(side=TOP,pady=10)

    payment = Label(root,text=PAYMENT.get(),font=("Bahnschrift SemiBold",12))
    payment.pack(side=TOP,pady=5)

    #button to save the data and go back to the store
    ok = Button(root,text="OK",font=("Bahnschrift SemiBold",12),
        command= lambda: [insert([NAME.get(),int(CONTACT.get()),EMAIL.get(),PAYMENT.get(),
            int(AMOUNT.get()),ITEMS.get(),DELIVERY.get(),ADDRESS.get()]),
            messagebox.showinfo(title="DO VISIT SOON",message="THANK YOU!!!"),dest(root)])
    ok.pack(side=TOP,pady=30)
    root.mainloop()

#function to save the data to the excel sheet
def insert(data):
    curr_row = sheet._current_row + 1
    for i in range(len(data)):
        sheet.cell(row=curr_row, column=i+1).value = data[i]

    wb.save("D:\I034, I042, I042, I049_DSA Project_MobiCart\Details.xlsx")

#the mainframe of the store
storeWindow = tk.Tk()
storeWindow.title("MobiStore")
storeWindow.geometry("500x500")
storeWindow.configure(background="light blue")

#Phone List
phone = ["Oneplus Nord CE","Oneplus Nord 2","Oneplus 8","Oneplus 8T","Oneplus 9R",
            "Oneplus 8 Pro","Oneplus 9","Oneplus 9 Pro","Apple Iphone SE ","Apple Iphone XR",
            "Apple Iphone 12","Apple Iphone 12 Pro","Apple Iphone 12 Pro Max",
            "Samsung Galaxy M11","Samsung Galaxy M21","Samsung Galaxy A12","Samsung Galaxy A31",
            "Samsung Galaxy M32","Samsung Galaxy A51","Samsung Galaxy A71","Samsung Galaxy S10e",
            "Samsung Galaxy S20+","Mi 11 Lite","Mi 11 X 5G","Mi 10T","Mi 11 Pro 5G",
            "Mi 11 Ultra 5G"] 

#Price List
price = [22999,27999,38999,39900,39900,
        48900,49999,64000,39999,54999,
        69900,109900,129650,
        9999,12999,13999,16999,
        18999,20999,27999,44999,
        60000,21999,29900,32999,39900,
        69000]

#tkinter variables to store data
NAME = StringVar()
CONTACT = StringVar()
EMAIL = StringVar()
PAYMENT = StringVar()
AMOUNT = StringVar()
ITEMS = StringVar()
DELIVERY = StringVar()
ADDRESS = StringVar()

headings()

viewStore(phone,price)
cart = ShoppingCart() 

storeWindow.mainloop()
